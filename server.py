from flask import Flask, render_template, request, redirect, send_file, jsonify
import sqlite3, os, base64, requests
import pytz
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

# ---------------- SUPABASE CONFIG ----------------

SUPABASE_URL = "https://odbkrbarwhhzfemqfbts.supabase.co"
SUPABASE_KEY = "sb_publishable_9xNlO3TyVXlolLDhjIuuFw_wqdHCTYh"

def send_to_supabase(emp_id,date,in_time,out_time):

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "employee_id": emp_id,
        "name": emp_id,
        "date": date + " " + in_time,
        "image_url": f"{emp_id}_{in_time}.jpg"
}
    }

    requests.post(
        f"{SUPABASE_URL}/rest/v1/attendance",
        json=payload,
        headers=headers
    )

# ---------------- LOCAL DATABASE ----------------

def db():
    return sqlite3.connect("database.db")

def create_tables():

    con = db()
    cur = con.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS employees(
    id TEXT,
    name TEXT,
    designation TEXT,
    location TEXT)
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS attendance(
    emp_id TEXT,
    date TEXT,
    in_time TEXT,
    out_time TEXT,
    gps TEXT)
    """)

    con.commit()

create_tables()

# ---------------- IMAGE SAVE ----------------

def save_image(data,name):

    if data=="":
        return

    img=data.split(",")[1]

    os.makedirs("static",exist_ok=True)

    with open("static/"+name,"wb") as f:

        f.write(base64.b64decode(img))

# ---------------- ROUTES ----------------

@app.route("/")
def login():
    return render_template("login.html")

@app.route("/dashboard")
def dashboard():

    con=db()
    cur=con.cursor()

    cur.execute("select * from employees")
    data=cur.fetchall()

    return render_template("dashboard.html", employees=data)

@app.route("/add_employee",methods=["GET","POST"])
def add_employee():

    if request.method=="POST":

        con=db()
        cur=con.cursor()

        cur.execute(
        "insert into employees values(?,?,?,?)",
        (
        request.form["id"],
        request.form["name"],
        request.form["designation"],
        request.form["location"]
        )
        )

        con.commit()

        return redirect("/dashboard")

    return render_template("add_employee.html")

@app.route("/attendance/<emp_id>", methods=["GET","POST"])
def attendance(emp_id):

    india = pytz.timezone("Asia/Kolkata")

    if request.method == "POST":

        type = request.form["type"]
        img = request.form["image"]
        gps = request.form.get("location","")

        now = datetime.now(india)

        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")

        con = db()
        cur = con.cursor()

        cur.execute(
        "SELECT * FROM attendance WHERE emp_id=? AND date=?",
        (emp_id,date)
        )

        row = cur.fetchone()

        # ---------- IN TIME ----------

        if type=="IN":

            if row:
                return redirect(f"/attendance/{emp_id}")

            cur.execute(
            "INSERT INTO attendance VALUES (?,?,?,?,?)",
            (emp_id,date,time,"",gps)
            )

            filename = f"{emp_id}_{time}.jpg"
            save_image(img, filename)

            # send to online database
            send_to_supabase(emp_id,date,time,"")

        # ---------- OUT TIME ----------

        else:

            if not row:
                return redirect(f"/attendance/{emp_id}")

            if row[3]!="":
                return redirect(f"/attendance/{emp_id}")

            cur.execute(
            """
            UPDATE attendance
            SET out_time=?, gps=?
            WHERE emp_id=? AND date=?
            """,
            (time,gps,emp_id,date)
            )

            save_image(img,f"{emp_id}_out.jpg")

            send_to_supabase(emp_id,date,row[2],time)

        con.commit()

        return redirect(f"/attendance/{emp_id}")

    # ---------- SHOW TIMES ----------

    con = db()
    cur = con.cursor()

    cur.execute(
    """
    SELECT in_time,out_time
    FROM attendance
    WHERE emp_id=?
    ORDER BY date DESC
    LIMIT 1
    """,
    (emp_id,)
    )

    row = cur.fetchone()

    in_time=""
    out_time=""

    if row:
        in_time=row[0]
        out_time=row[1]

    return render_template(
    "attendance.html",
    emp_id=emp_id,
    in_time=in_time,
    out_time=out_time
    )

# ---------------- REPORT ----------------

@app.route("/report",methods=["GET","POST"])
def report():

    data=[]

    if request.method=="POST":

        f=request.form["from"]
        t=request.form["to"]

        con=db()
        cur=con.cursor()

        cur.execute(
        """
        select * from attendance
        where date between ? and ?
        """,
        (f,t)
        )

        data=cur.fetchall()

    return render_template("report.html",data=data)

# ---------------- EXCEL ----------------

@app.route("/excel")
def excel():

    con=db()
    cur=con.cursor()

    cur.execute("select * from attendance")
    rows=cur.fetchall()

    wb=Workbook()
    ws=wb.active

    ws.append(["Emp","Date","Status"])

    green=PatternFill(start_color="00FF00",fill_type="solid")
    yellow=PatternFill(start_color="FFFF00",fill_type="solid")
    red=PatternFill(start_color="FF0000",fill_type="solid")

    for r in rows:

        if r[3]=="":
            status="A"
            color=red

        else:

            t1=datetime.strptime(r[2],"%H:%M:%S")
            t2=datetime.strptime(r[3],"%H:%M:%S")

            hrs=(t2-t1).seconds/3600

            if hrs>=8:
                status="P"
                color=green

            elif hrs>=4:
                status="PH"
                color=yellow

            else:
                status="A"
                color=red

        ws.append([r[0],r[1],status])
        ws.cell(ws.max_row,3).fill=color

    wb.save("report.xlsx")

    return send_file("report.xlsx",as_attachment=True)

# ---------------- EMPLOYEE ----------------

@app.route("/delete_employee/<id>")
def delete_employee(id):

    con=db()
    cur=con.cursor()

    cur.execute("delete from employees where id=?",(id,))
    con.commit()

    return redirect("/dashboard")

@app.route("/edit_employee/<id>",methods=["GET","POST"])
def edit_employee(id):

    con=db()
    cur=con.cursor()

    if request.method=="POST":

        cur.execute(
        """
        update employees
        set name=?,designation=?,location=?
        where id=?
        """,
        (
        request.form["name"],
        request.form["designation"],
        request.form["location"],
        id
        )
        )

        con.commit()

        return redirect("/dashboard")

    cur.execute("select * from employees where id=?",(id,))
    emp=cur.fetchone()

    return render_template("edit_employee.html",emp=emp)

# ---------------- RUN ----------------

if __name__=="__main__":

    port=int(os.environ.get("PORT",5000))

    app.run(host="0.0.0.0",port=port)
